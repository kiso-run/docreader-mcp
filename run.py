"""tool-docreader — extract text from PDF, DOCX, CSV, XLSX, and plain text files.

Subprocess contract (same as all kiso tools):
  stdin:  JSON {args, session, workspace, session_secrets, plan_outputs}
  stdout: result text on success
  stderr: error description on failure
  exit 0: success, exit 1: failure
"""
from __future__ import annotations

import csv
import io
import json
import signal
import sys
from pathlib import Path

signal.signal(signal.SIGTERM, lambda *_: sys.exit(0))

# Maximum characters to output (prevent memory exhaustion on huge files).
_MAX_OUTPUT_CHARS = 100_000

# Plain text extensions (read as-is).
_TEXT_EXTENSIONS = frozenset({
    ".txt", ".md", ".rst", ".log", ".json", ".yaml", ".yml",
    ".toml", ".ini", ".cfg", ".conf", ".sh", ".bash", ".py",
    ".js", ".ts", ".html", ".htm", ".xml", ".css", ".sql",
})


def main() -> None:
    data = json.load(sys.stdin)
    args = data.get("args", {})
    workspace = data.get("workspace", ".")

    action = args.get("action", "read")

    try:
        if action == "list":
            result = do_list(workspace)
        elif action == "info":
            result = do_info(workspace, args)
        elif action == "read":
            result = do_read(workspace, args)
        else:
            print(f"Unknown action: {action}", file=sys.stderr)
            sys.exit(1)
    except FileNotFoundError as e:
        print(f"File not found: {e}", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)

    print(result)


# ---------------------------------------------------------------------------
# Actions
# ---------------------------------------------------------------------------


def do_list(workspace: str) -> str:
    """List files in the uploads/ directory."""
    uploads = Path(workspace) / "uploads"
    if not uploads.is_dir():
        return "No uploads/ directory found."
    files = sorted(f for f in uploads.rglob("*") if f.is_file())
    if not files:
        return "uploads/ directory is empty."
    lines = [f"Files in uploads/ ({len(files)}):"]
    for f in files:
        rel = f.relative_to(uploads)
        size = f.stat().st_size
        lines.append(f"  {rel} ({_format_size(size)})")
    return "\n".join(lines)


def do_info(workspace: str, args: dict) -> str:
    """Get file metadata without full extraction."""
    file_path = _resolve_path(workspace, args)
    ext = file_path.suffix.lower()
    size = file_path.stat().st_size
    lines = [
        f"File: {file_path.name}",
        f"Size: {_format_size(size)}",
        f"Format: {ext}",
    ]
    if ext == ".pdf":
        from pypdf import PdfReader
        reader = PdfReader(str(file_path))
        lines.append(f"Pages: {len(reader.pages)}")
    elif ext == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(str(file_path), read_only=True)
        lines.append(f"Sheets: {', '.join(wb.sheetnames)}")
        wb.close()
    elif ext == ".csv":
        with open(file_path, newline="", encoding="utf-8", errors="replace") as f:
            row_count = sum(1 for _ in csv.reader(f))
        lines.append(f"Rows: {row_count}")
    return "\n".join(lines)


def do_read(workspace: str, args: dict) -> str:
    """Extract text content from a file."""
    file_path = _resolve_path(workspace, args)
    ext = file_path.suffix.lower()

    if ext == ".pdf":
        text = _read_pdf(file_path, args.get("pages"))
    elif ext == ".docx":
        text = _read_docx(file_path)
    elif ext == ".xlsx":
        text = _read_xlsx(file_path)
    elif ext == ".csv":
        text = _read_csv(file_path)
    elif ext in _TEXT_EXTENSIONS or _is_likely_text(file_path):
        text = _read_text(file_path)
    else:
        return f"Unsupported file format: {ext}. Supported: PDF, DOCX, XLSX, CSV, and plain text."

    if len(text) > _MAX_OUTPUT_CHARS:
        text = text[:_MAX_OUTPUT_CHARS] + f"\n\n... (truncated at {_MAX_OUTPUT_CHARS} characters)"
    return text


# ---------------------------------------------------------------------------
# Format readers
# ---------------------------------------------------------------------------


def _read_pdf(path: Path, pages_arg: str | None = None) -> str:
    """Extract text from a PDF file, optionally only specific pages."""
    from pypdf import PdfReader
    reader = PdfReader(str(path))
    total_pages = len(reader.pages)

    if pages_arg:
        indices = _parse_page_ranges(pages_arg, total_pages)
    else:
        indices = range(total_pages)

    parts: list[str] = []
    for i in indices:
        text = reader.pages[i].extract_text() or ""
        if text.strip():
            parts.append(f"--- Page {i + 1} ---\n{text.strip()}")
    if not parts:
        return f"PDF has {total_pages} pages but no extractable text."
    return "\n\n".join(parts)


def _read_docx(path: Path) -> str:
    """Extract text from a DOCX file."""
    from docx import Document
    doc = Document(str(path))
    parts = [p.text for p in doc.paragraphs if p.text.strip()]
    if not parts:
        return "DOCX file has no text content."
    return "\n\n".join(parts)


def _read_xlsx(path: Path) -> str:
    """Extract text from an XLSX file (all sheets)."""
    from openpyxl import load_workbook
    wb = load_workbook(str(path), read_only=True, data_only=True)
    parts: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[str] = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                rows.append("\t".join(cells))
        if rows:
            parts.append(f"--- Sheet: {sheet_name} ---\n" + "\n".join(rows))
    wb.close()
    if not parts:
        return "XLSX file has no data."
    return "\n\n".join(parts)


def _read_csv(path: Path) -> str:
    """Extract text from a CSV file."""
    with open(path, newline="", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f)
        rows = ["\t".join(row) for row in reader]
    if not rows:
        return "CSV file is empty."
    return "\n".join(rows)


def _read_text(path: Path) -> str:
    """Read a plain text file."""
    return path.read_text(encoding="utf-8", errors="replace")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _resolve_path(workspace: str, args: dict) -> Path:
    """Resolve file_path arg to an absolute Path."""
    file_path = args.get("file_path")
    if not file_path:
        raise ValueError("file_path argument is required for read/info actions")
    resolved = (Path(workspace) / file_path).resolve()
    # Path traversal guard
    ws_resolved = Path(workspace).resolve()
    if not str(resolved).startswith(str(ws_resolved)):
        raise ValueError(f"Path traversal denied: {file_path}")
    if not resolved.is_file():
        raise FileNotFoundError(resolved.name)
    return resolved


def _parse_page_ranges(spec: str, total: int) -> list[int]:
    """Parse page range spec like '1-5,7,10-12' into zero-based indices."""
    indices: list[int] = []
    for part in spec.split(","):
        part = part.strip()
        if "-" in part:
            start, end = part.split("-", 1)
            start_i = max(0, int(start) - 1)
            end_i = min(total, int(end))
            indices.extend(range(start_i, end_i))
        else:
            i = int(part) - 1
            if 0 <= i < total:
                indices.append(i)
    return sorted(set(indices))


def _format_size(size: int) -> str:
    """Format byte size as human-readable string."""
    if size < 1024:
        return f"{size} B"
    if size < 1024 * 1024:
        return f"{size / 1024:.1f} KB"
    return f"{size / (1024 * 1024):.1f} MB"


def _is_likely_text(path: Path) -> bool:
    """Heuristic: check if the first 512 bytes look like text."""
    try:
        sample = path.read_bytes()[:512]
        # If mostly printable ASCII + whitespace, treat as text
        text_chars = sum(1 for b in sample if 32 <= b < 127 or b in (9, 10, 13))
        return len(sample) > 0 and text_chars / len(sample) > 0.85
    except Exception:
        return False


if __name__ == "__main__":
    main()
